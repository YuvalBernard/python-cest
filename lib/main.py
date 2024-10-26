"""
Module that defines the app structure
"""

import sys

import openpyxl
import pandas as pd
from PyQt6.QtCore import QSize, Qt
from PyQt6.QtGui import QFont, QPixmap
from PyQt6.QtWidgets import (
    QApplication,
    QCheckBox,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWizard,
    QWizardPage,
)


class IntroPage(QWizardPage):
    def __init__(self):
        super().__init__()

        self.setTitle("Welcome!")
        self.setPixmap(QWizard.WizardPixmap.WatermarkPixmap, QPixmap("lib/images/watermark.jpg"))
        self.setPixmap(QWizard.WizardPixmap.LogoPixmap, QPixmap("lib/images/logo.png"))

        label = QLabel(
            "This Wizard will walk you through all the neccessary steps "
            "to fit Z-spectra to Bloch-McConnell equations. "
            "You may chose a solver and a fitting method."
        )
        label.setWordWrap(True)

        layout = QVBoxLayout(self)
        layout.addWidget(label)


class DataPage(QWizardPage):
    def __init__(self):
        super().__init__()

        self.setTitle("Select Data")
        page_label = QLabel(
            "Select .xlsx file with data to fit. Each column should be attributed to a different saturation power."
        )
        page_label.setWordWrap(True)

        layout = QVBoxLayout(self)
        layout.addWidget(page_label)

        open_button = QPushButton("Open file")
        open_button.clicked.connect(self.get_filename)
        self.label = QLabel("None selected.")

        file_select_layout = QHBoxLayout()
        file_select_layout.addWidget(open_button)
        file_select_layout.addWidget(self.label)
        layout.addLayout(file_select_layout)

        self.table = QTableWidget()
        layout.addWidget(self.table)

        save_button = QPushButton("Save table")
        # Clicking the button is a mandatory operation to proceed.
        # save_button.setCheckable(True)
        # self.registerField("*", save_button)
        save_button.clicked.connect(self.save_table)
        layout.addWidget(save_button)

    def get_filename(self):
        self.filename, extension = QFileDialog.getOpenFileName(self)
        self.label.setText(self.filename)
        if self.filename:
            self.update_table()

    def update_table(self):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook.active
        self.table.setRowCount(sheet.max_row)
        self.table.setColumnCount(sheet.max_column)
        list_vals = list(sheet.values)
        self.table.setHorizontalHeaderLabels(list_vals[0])
        row_index = 0
        for value_tuple in list_vals[1:]:
            col_index = 0
            for value in value_tuple:
                self.table.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1
        self.table.update()

    def save_table(self):
        # Save updated table to DataFrame for later use
        self.wizard().dataframe = self.write_qtable_to_df(self.table)

    @staticmethod
    def write_qtable_to_df(table: QTableWidget):
        col_count = table.columnCount()
        row_count = table.rowCount()
        headers = [str(table.horizontalHeaderItem(i).text()) for i in range(col_count)]

        # df indexing is slow, so use lists
        df_list = []
        for row in range(row_count):
            df_list2 = []
            for col in range(col_count):
                table_item = table.item(row, col)
                df_list2.append("" if table_item is None else str(table_item.text()))
            df_list.append(df_list2)

        df = pd.DataFrame(df_list, columns=headers)

        return df


@staticmethod
def makeBold(widget):
    f = widget.font()
    f.setWeight(QFont.Weight.DemiBold)
    widget.setFont(f)
    return widget


class ConstantsGroup(QGroupBox):
    def __init__(self, parent: QWizardPage):
        super().__init__(parent)

        self.setTitle("Constants")
        layout = QGridLayout(self)

        b0_label = QLabel("B₀ (T)")
        gamma_label = QLabel("γ (MHz/T)")
        tp_label = QLabel("tₚ (s)")
        b1_label = QLabel("B₁ (µT)")

        b0_desc = QLabel("External magnetic field strength")
        gamma_desc = QLabel("Gyromagnetic ratio")
        tp_desc = QLabel("Saturation pulse duration")
        b1_desc = QLabel("Saturation pulse amplitude")

        b0_entry = QLineEdit()
        gamma_entry = QLineEdit()
        tp_entry = QLineEdit()
        b1_entry = QLineEdit()

        # Make values mandatory
        parent.registerField("b0", b0_entry)
        parent.registerField("gamma", gamma_entry)
        parent.registerField("tp", tp_entry)
        parent.registerField("b1", b1_entry)

        layout.addWidget(makeBold(QLabel("Parameter")), 0, 0)
        layout.addWidget(makeBold(QLabel("Value")), 0, 1)
        layout.addWidget(makeBold(QLabel("Description")), 0, 2)
        layout.addWidget(b0_label, 1, 0)
        layout.addWidget(b0_entry, 1, 1)
        layout.addWidget(b0_desc, 1, 2)
        layout.addWidget(gamma_label, 2, 0)
        layout.addWidget(gamma_entry, 2, 1)
        layout.addWidget(gamma_desc, 2, 2)
        layout.addWidget(tp_label, 3, 0)
        layout.addWidget(tp_entry, 3, 1)
        layout.addWidget(tp_desc, 3, 2)
        layout.addWidget(b1_label, 4, 0)
        layout.addWidget(b1_entry, 4, 1)
        layout.addWidget(b1_desc, 4, 2)


class VariablesGroup(QGroupBox):
    def __init__(self, parent: QWizardPage):
        super().__init__(parent)

        self.setTitle("Variables")
        layout = QGridLayout(self)

        R1a_label = QLabel("R1a (Hz)")
        R2a_label = QLabel("R2a (Hz)")
        dwa_label = QLabel("dwa (ppm)")
        R1b_label = QLabel("R1b (Hz)")
        R2b_label = QLabel("R2b (Hz)")
        kb_label = QLabel("kb (Hz)")
        fb_label = QLabel("fb (Hz)")
        dwb_label = QLabel("dwb (ppm)")

        R1a_desc = QLabel("Pool A longitudinal relaxation rate")
        R2a_desc = QLabel("Pool A transverse relaxation rate")
        dwa_desc = QLabel("Larmor frequency of pool A relative to itself. Optimally zero")
        R1b_desc = QLabel("Pool B longitudinal relaxation rate")
        R2b_desc = QLabel("Pool B transverse relaxation rate")
        kb_desc = QLabel("Pool B forward exchange rate")
        fb_desc = QLabel("Pool B equilibrium magnetization; fraction relative to pool A")
        dwb_desc = QLabel("Larmor frequency of pool B relative to pool A")

        R1a_vary = QCheckBox()
        R2a_vary = QCheckBox()
        dwa_vary = QCheckBox()
        R1b_vary = QCheckBox()
        R2b_vary = QCheckBox()
        kb_vary = QCheckBox()
        fb_vary = QCheckBox()
        dwb_vary = QCheckBox()

        R1a_val = QLineEdit()
        R2a_val = QLineEdit()
        dwa_val = QLineEdit()
        R1b_val = QLineEdit()
        R2b_val = QLineEdit()
        kb_val = QLineEdit()
        fb_val = QLineEdit()
        dwb_val = QLineEdit()

        R1a_min = QLineEdit()
        R2a_min = QLineEdit()
        dwa_min = QLineEdit()
        R1b_min = QLineEdit()
        R2b_min = QLineEdit()
        kb_min = QLineEdit()
        fb_min = QLineEdit()
        dwb_min = QLineEdit()

        R1a_max = QLineEdit()
        R2a_max = QLineEdit()
        dwa_max = QLineEdit()
        R1b_max = QLineEdit()
        R2b_max = QLineEdit()
        kb_max = QLineEdit()
        fb_max = QLineEdit()
        dwb_max = QLineEdit()

        # Create registerFields to track values
        for checkbox, min, max, val, name_min, name_max, name_val in zip(
            [R1a_vary, R2a_vary, dwa_vary, R1b_vary, R2b_vary, kb_vary, fb_vary, dwb_vary],
            [R1a_min, R2a_min, dwa_min, R1b_min, R2b_min, kb_min, fb_min, dwb_min],
            [R1a_max, R2a_max, dwa_max, R1b_max, R2b_max, kb_max, fb_max, dwb_max],
            [R1a_val, R2a_val, dwa_val, R1b_val, R2b_val, kb_val, fb_val, dwb_val],
            ["R1a_min", "R2a_min", "dwa_min", "R1b_min", "R2b_min", "kb_min", "fb_min", "dwb_min"],
            ["R1a_max", "R2a_max", "dwa_max", "R1b_max", "R2b_max", "kb_max", "fb_max", "dwb_max"],
            ["R1a_val", "R2a_val", "dwa_val", "R1b_val", "R2b_val", "kb_val", "fb_val", "dwb_val"],
        ):
            if checkbox.checkState() == Qt.CheckState.Checked:
                min.setEnabled(True)
                max.setEnabled(True)
                parent.registerField(name_min, min)
                parent.registerField(name_max, max)

            else:
                min.setDisabled(True)
                max.setDisabled(True)
                parent.registerField(name_val, val)

            checkbox.toggled.connect(min.setEnabled)
            checkbox.toggled.connect(max.setEnabled)
            checkbox.toggled.connect(val.setDisabled)

        layout.addWidget(makeBold(QLabel("Parameter")), 0, 0)
        layout.addWidget(makeBold(QLabel("Vary?\n(Y/N)")), 0, 1)
        layout.addWidget(makeBold(QLabel("Min")), 0, 2)
        layout.addWidget(makeBold(QLabel("Max")), 0, 3)
        layout.addWidget(makeBold(QLabel("Value")), 0, 4)
        layout.addWidget(makeBold(QLabel("Description")), 0, 5)

        layout.addWidget(R1a_label, 1, 0)
        layout.addWidget(R1a_vary, 1, 1)
        layout.addWidget(R1a_min, 1, 2)
        layout.addWidget(R1a_max, 1, 3)
        layout.addWidget(R1a_val, 1, 4)
        layout.addWidget(R1a_desc, 1, 5)
        layout.addWidget(R2a_label, 2, 0)
        layout.addWidget(R2a_vary, 2, 1)
        layout.addWidget(R2a_min, 2, 2)
        layout.addWidget(R2a_max, 2, 3)
        layout.addWidget(R2a_val, 2, 4)
        layout.addWidget(R2a_desc, 2, 5)
        layout.addWidget(dwa_label, 3, 0)
        layout.addWidget(dwa_vary, 3, 1)
        layout.addWidget(dwa_min, 3, 2)
        layout.addWidget(dwa_max, 3, 3)
        layout.addWidget(dwa_val, 3, 4)
        layout.addWidget(dwa_desc, 3, 5)
        layout.addWidget(R1b_label, 4, 0)
        layout.addWidget(R1b_vary, 4, 1)
        layout.addWidget(R1b_min, 4, 2)
        layout.addWidget(R1b_max, 4, 3)
        layout.addWidget(R1b_val, 4, 4)
        layout.addWidget(R1b_desc, 4, 5)
        layout.addWidget(R2b_label, 5, 0)
        layout.addWidget(R2b_vary, 5, 1)
        layout.addWidget(R2b_min, 5, 2)
        layout.addWidget(R2b_max, 5, 3)
        layout.addWidget(R2b_val, 5, 4)
        layout.addWidget(R2b_desc, 5, 5)
        layout.addWidget(kb_label, 6, 0)
        layout.addWidget(kb_vary, 6, 1)
        layout.addWidget(kb_min, 6, 2)
        layout.addWidget(kb_max, 6, 3)
        layout.addWidget(kb_val, 6, 4)
        layout.addWidget(kb_desc, 6, 5)
        layout.addWidget(fb_label, 7, 0)
        layout.addWidget(fb_vary, 7, 1)
        layout.addWidget(fb_min, 7, 2)
        layout.addWidget(fb_max, 7, 3)
        layout.addWidget(fb_val, 7, 4)
        layout.addWidget(fb_desc, 7, 5)
        layout.addWidget(dwb_label, 8, 0)
        layout.addWidget(dwb_vary, 8, 1)
        layout.addWidget(dwb_min, 8, 2)
        layout.addWidget(dwb_max, 8, 3)
        layout.addWidget(dwb_val, 8, 4)
        layout.addWidget(dwb_desc, 8, 5)


class ModelPage(QWizardPage):
    def __init__(self):
        super().__init__()

        self.setTitle("Configure Model Parameters")
        layout = QVBoxLayout(self)
        layout.addWidget(ConstantsGroup(self))
        layout.addWidget(VariablesGroup(self))


class WizardApp(QWizard):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.setWindowTitle("Bloch-McConnellizer")
        self.setMinimumSize(QSize(845, 585))
        self.addPage(IntroPage())
        self.addPage(DataPage())
        self.addPage(ModelPage())


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = WizardApp()
    window.show()
    sys.exit(app.exec())
