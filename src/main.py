"""
Module that defines the app structure
"""

import sys

import lmfit
import numpy as np
import openpyxl
import pandas as pd
from PyQt6.QtCore import QSize, Qt
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QItemDelegate,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QStackedLayout,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QWizard,
    QWizardPage,
)

from fit_spectra import bayesian_mcmc, bayesian_vi, least_squares
from simulate_spectra import batch_gen_spectrum_analytical, batch_gen_spectrum_numerical, batch_gen_spectrum_symbolic


class IntroPage(QWizardPage):
    def __init__(self):
        super().__init__()

        self.setTitle("Welcome!")
        # self.setPixmap(QWizard.WizardPixmap.WatermarkPixmap, QPixmap("src/images/watermark.jpg"))
        # self.setPixmap(QWizard.WizardPixmap.LogoPixmap, QPixmap("src/images/logo.png"))

        label = QLabel(
            "This Wizard will walk you through all the neccessary steps "
            "to fit Z-spectra to Bloch-McConnell equations. "
            "You may chose a solver and a fitting method."
        )
        label.setWordWrap(True)

        layout = QVBoxLayout(self)
        layout.addWidget(label)


class FloatDelegate(QItemDelegate):
    def __init__(self, decimals, parent=None):
        QItemDelegate.__init__(self, parent=parent)
        self.nDecimals = decimals

    def paint(self, painter, option, index):
        value = index.model().data(index, Qt.ItemDataRole.EditRole)
        try:
            number = float(value)
            painter.drawText(option.rect, Qt.AlignmentFlag.AlignLeft, "{:.{}f}".format(number, self.nDecimals))
        except:
            QItemDelegate.paint(self, painter, option, index)


class DataPage(QWizardPage):
    def __init__(self):
        super().__init__()

        self.setTitle("Select Data")
        page_label = QLabel(
            "Select .xlsx file with data to fit.\n"
            "Each column should be attributed to a different saturation power (in µT)."
        )
        page_label.setWordWrap(True)

        layout = QVBoxLayout(self)
        layout.addWidget(page_label)

        open_button = QPushButton("Open file")
        open_button.clicked.connect(self.get_filename)
        self.label = QLabel("None selected.")

        file_select_layout = QHBoxLayout()
        file_select_layout.addWidget(open_button)
        file_select_layout.addWidget(self.label)
        layout.addLayout(file_select_layout)

        self.table = QTableWidget()
        self.table.setItemDelegate(FloatDelegate(3))
        layout.addWidget(self.table)
        self.table.itemChanged.connect(self.save_table)

    def get_filename(self):
        self.filename, extension = QFileDialog.getOpenFileName(self)
        self.label.setText(self.filename)
        if self.filename:
            self.update_table()
            self.save_table()

    def update_table(self):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook.active
        self.table.setRowCount(sheet.max_row)
        self.table.setColumnCount(sheet.max_column)
        list_vals = list(sheet.values)
        self.table.setHorizontalHeaderLabels(list_vals[0])
        row_index = 0
        for value_tuple in list_vals[1:]:
            col_index = 0
            for value in value_tuple:
                self.table.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1
        self.table.update()

    def save_table(self):
        # Save updated table to DataFrame for later use
        self.wizard().dataframe = write_qtable_to_df(self.table)


class ConstantsGroup(QGroupBox):
    def __init__(self, parent: QWizardPage):
        super().__init__(parent)

        self.setTitle("Constants")
        layout = QGridLayout(self)

        b0_label = QLabel("B₀ (T)")
        gamma_label = QLabel("γ (MHz/T)")
        tp_label = QLabel("tₚ (s)")
        b1_label = QLabel("B₁ (µT)")

        b0_desc = QLabel("External magnetic field strength")
        gamma_desc = QLabel("Gyromagnetic ratio")
        tp_desc = QLabel("Saturation pulse duration")
        b1_desc = QLabel("Saturation pulse amplitude\n(comma separated list)")

        b0_entry = QLineEdit()
        gamma_entry = QLineEdit()
        tp_entry = QLineEdit()
        self.b1_entry = QLineEdit(placeholderText="e.g. 1.0, 3.0, 5.0")

        parent.registerField("b0", b0_entry)
        parent.registerField("gamma", gamma_entry)
        parent.registerField("tp", tp_entry)
        parent.registerField("b1", self.b1_entry)

        layout.addWidget(makeBold(QLabel("Parameter")), 0, 0)
        layout.addWidget(makeBold(QLabel("Value")), 0, 1)
        layout.addWidget(makeBold(QLabel("Description")), 0, 2)
        layout.addWidget(b0_label, 1, 0, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(b0_entry, 1, 1, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(b0_desc, 1, 2, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(gamma_label, 2, 0, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(gamma_entry, 2, 1, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(gamma_desc, 2, 2, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(tp_label, 3, 0, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(tp_entry, 3, 1, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(tp_desc, 3, 2, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(b1_label, 4, 0, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(self.b1_entry, 4, 1, alignment=Qt.AlignmentFlag.AlignTop)
        layout.addWidget(b1_desc, 4, 2, alignment=Qt.AlignmentFlag.AlignTop)

        button = QPushButton("Fill B₁ from table")
        button.clicked.connect(lambda: self.fill_in_b1_list(parent))
        layout.addWidget(button, 5, 0, 1, -1)

        # For testing purposes:
        b0_entry.setText("9.4")
        gamma_entry.setText("16.546")
        tp_entry.setText("2")

    def fill_in_b1_list(self, parent):
        if hasattr(parent.wizard(), "dataframe"):
            df = parent.wizard().dataframe
            # if the excel sheet headers are in µT:
            if df.columns[1:].str.contains("T").any():
                b1_list = list(df.columns[1:].str.extract(r"([-+]?\d*\.?\d+)", expand=False))
                self.b1_entry.setText(", ".join([f"{float(b1):.2f}" for b1 in b1_list]))

        else:
            QMessageBox.warning(self, "Warning", "Please select data excel sheet on previous page.")


class VariablesGroup(QGroupBox):
    def __init__(self, parent: QWizardPage):
        super().__init__(parent)

        self.setTitle("Variables")
        layout = QGridLayout(self)

        R1a_label = QLabel("R1a (Hz)")
        R2a_label = QLabel("R2a (Hz)")
        dwa_label = QLabel("dwa (ppm)")
        R1b_label = QLabel("R1b (Hz)")
        R2b_label = QLabel("R2b (Hz)")
        kb_label = QLabel("kb (Hz)")
        fb_label = QLabel("fb (Hz)")
        dwb_label = QLabel("dwb (ppm)")

        R1a_desc = QLabel("Pool A longitudinal relaxation rate")
        R2a_desc = QLabel("Pool A transverse relaxation rate")
        dwa_desc = QLabel("Larmor frequency of pool A relative to itself. Optimally zero")
        R1b_desc = QLabel("Pool B longitudinal relaxation rate")
        R2b_desc = QLabel("Pool B transverse relaxation rate")
        kb_desc = QLabel("Pool B forward exchange rate")
        fb_desc = QLabel("Pool B equilibrium magnetization; fraction relative to pool A")
        dwb_desc = QLabel("Larmor frequency of pool B relative to pool A")

        R1a_vary = QComboBox()
        R2a_vary = QComboBox()
        dwa_vary = QComboBox()
        R1b_vary = QComboBox()
        R2b_vary = QComboBox()
        kb_vary = QComboBox()
        fb_vary = QComboBox()
        dwb_vary = QComboBox()

        for combo_box, name in zip(
            [R1a_vary, R2a_vary, dwa_vary, R1b_vary, R2b_vary, kb_vary, fb_vary, dwb_vary],
            [
                "R1a_vary",
                "R2a_vary",
                "dwa_vary",
                "R1b_vary",
                "R2b_vary",
                "kb_vary",
                "fb_vary",
                "dwb_vary",
            ],
        ):
            combo_box.addItems(["Vary", "Static"])
            parent.registerField(name, combo_box)

        R1a_val = QLineEdit()
        R2a_val = QLineEdit()
        dwa_val = QLineEdit()
        R1b_val = QLineEdit()
        R2b_val = QLineEdit()
        kb_val = QLineEdit()
        fb_val = QLineEdit()
        dwb_val = QLineEdit()

        R1a_init_val = QLineEdit()
        R2a_init_val = QLineEdit()
        dwa_init_val = QLineEdit()
        R1b_init_val = QLineEdit()
        R2b_init_val = QLineEdit()
        kb_init_val = QLineEdit()
        fb_init_val = QLineEdit()
        dwb_init_val = QLineEdit()

        R1a_min = QLineEdit()
        R2a_min = QLineEdit()
        dwa_min = QLineEdit()
        R1b_min = QLineEdit()
        R2b_min = QLineEdit()
        kb_min = QLineEdit()
        fb_min = QLineEdit()
        dwb_min = QLineEdit()

        R1a_max = QLineEdit()
        R2a_max = QLineEdit()
        dwa_max = QLineEdit()
        R1b_max = QLineEdit()
        R2b_max = QLineEdit()
        kb_max = QLineEdit()
        fb_max = QLineEdit()
        dwb_max = QLineEdit()

        layout.addWidget(makeBold(QLabel("Parameter")), 0, 0)
        layout.addWidget(makeBold(QLabel("State")), 0, 1)
        layout.addWidget(makeBold(QLabel("Min")), 0, 2)
        layout.addWidget(makeBold(QLabel("Max")), 0, 3)
        layout.addWidget(makeBold(QLabel("Init Value")), 0, 4)
        layout.addWidget(makeBold(QLabel("Static Value")), 0, 5)
        layout.addWidget(makeBold(QLabel("Description")), 0, 6)

        # Create registerFields to track values
        for i, (
            label,
            combo_box,
            min,
            max,
            val,
            init_val,
            desc,
            name_min,
            name_max,
            name_val,
            name_init_val,
        ) in enumerate(
            zip(
                [
                    R1a_label,
                    R2a_label,
                    dwa_label,
                    R1b_label,
                    R2b_label,
                    kb_label,
                    fb_label,
                    dwb_label,
                ],
                [R1a_vary, R2a_vary, dwa_vary, R1b_vary, R2b_vary, kb_vary, fb_vary, dwb_vary],
                [R1a_min, R2a_min, dwa_min, R1b_min, R2b_min, kb_min, fb_min, dwb_min],
                [R1a_max, R2a_max, dwa_max, R1b_max, R2b_max, kb_max, fb_max, dwb_max],
                [R1a_val, R2a_val, dwa_val, R1b_val, R2b_val, kb_val, fb_val, dwb_val],
                [
                    R1a_init_val,
                    R2a_init_val,
                    dwa_init_val,
                    R1b_init_val,
                    R2b_init_val,
                    kb_init_val,
                    fb_init_val,
                    dwb_init_val,
                ],
                [R1a_desc, R2a_desc, dwa_desc, R1b_desc, R2b_desc, kb_desc, fb_desc, dwb_desc],
                [
                    "R1a_min",
                    "R2a_min",
                    "dwa_min",
                    "R1b_min",
                    "R2b_min",
                    "kb_min",
                    "fb_min",
                    "dwb_min",
                ],
                [
                    "R1a_max",
                    "R2a_max",
                    "dwa_max",
                    "R1b_max",
                    "R2b_max",
                    "kb_max",
                    "fb_max",
                    "dwb_max",
                ],
                [
                    "R1a_val",
                    "R2a_val",
                    "dwa_val",
                    "R1b_val",
                    "R2b_val",
                    "kb_val",
                    "fb_val",
                    "dwb_val",
                ],
                [
                    "R1a_init_val",
                    "R2a_init_val",
                    "dwa_init_val",
                    "R1b_init_val",
                    "R2b_init_val",
                    "kb_init_val",
                    "fb_init_val",
                    "dwb_init_val",
                ],
            )
        ):
            combo_box.currentIndexChanged.connect(min.setDisabled)
            combo_box.currentIndexChanged.connect(max.setDisabled)
            combo_box.currentIndexChanged.connect(init_val.setDisabled)
            combo_box.currentIndexChanged.connect(val.setEnabled)

            parent.registerField(name_min, min)
            parent.registerField(name_max, max)
            parent.registerField(name_init_val, init_val)
            parent.registerField(name_val, val)

            if combo_box.currentText() == "Vary":
                min.setEnabled(True)
                max.setEnabled(True)
                init_val.setEnabled(True)
                val.setDisabled(True)
            else:
                min.setDisabled(True)
                max.setDisabled(True)
                init_val.setDisabled(True)
                val.setEnabled(True)

            j = i + 1
            layout.addWidget(label, j, 0)
            layout.addWidget(combo_box, j, 1)
            layout.addWidget(min, j, 2)
            layout.addWidget(max, j, 3)
            layout.addWidget(init_val, j, 4)
            layout.addWidget(val, j, 5)
            layout.addWidget(desc, j, 6)

        # For testing purposes:
        R1a_val.setText("8")
        R2a_val.setText("380")
        dwa_val.setText("0")
        R1b_min.setText("0.1")
        R1b_max.setText("10")
        R1b_init_val.setText("5")
        R2b_min.setText("1e4")
        R2b_max.setText("1e5")
        R2b_init_val.setText("5e4")
        kb_min.setText("100")
        kb_max.setText("500")
        kb_init_val.setText("150")
        fb_min.setText("1e-4")
        fb_max.setText("5e-2")
        fb_init_val.setText("1e-3")
        dwb_min.setText("-265")
        dwb_max.setText("-255")
        dwb_init_val.setText("-260")


class SolverGroup(QGroupBox):
    def __init__(self, parent: QWizardPage):
        super().__init__(parent)

        self.setTitle("Solver and Fitting Method")
        layout = QVBoxLayout(self)

        solver_list = QComboBox()
        solver_list.addItems(["Symbolic", "Analytical", "Numerical"])
        parent.registerField("solver", solver_list)

        fitter_list = QComboBox()
        fitter_list.addItems(["Bayesian, MCMC", "Bayesian, ADVI", "Nonlinear Least Squares"])
        parent.registerField("fitting_method", fitter_list)

        selection_layout = QFormLayout()
        selection_layout.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.FieldsStayAtSizeHint)
        selection_layout.addRow("Solver:", solver_list)
        selection_layout.addRow("Fitting method:", fitter_list)
        layout.addLayout(selection_layout)
        layout.addSpacing(20)
        layout.addWidget(makeBold(QLabel("Configure fitting method")))

        config_page = QStackedLayout()
        layout.addLayout(config_page)
        fitter_list.currentIndexChanged.connect(config_page.setCurrentIndex)

        # Configure MCMC widget
        bayesian_mcmc_container = QWidget()
        bayesian_mcmc_layout = QFormLayout(bayesian_mcmc_container)
        bayesian_mcmc_layout.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.FieldsStayAtSizeHint)

        warmup_entry = QLineEdit("1000")
        samples_entry = QLineEdit("1000")
        chains_entry = QLineEdit("4")

        parent.registerField("num_warmup", warmup_entry)
        parent.registerField("num_samples", samples_entry)
        parent.registerField("num_chains", chains_entry)

        bayesian_mcmc_layout.addRow("Number of warmup samples", warmup_entry)
        bayesian_mcmc_layout.addRow("Number of posterior samples", samples_entry)
        bayesian_mcmc_layout.addRow("Number of chains", chains_entry)
        config_page.addWidget(bayesian_mcmc_container)

        # Configure ADVI widget
        bayesian_vi_container = QWidget()
        bayesian_vi_layout = QFormLayout(bayesian_vi_container)
        bayesian_vi_layout.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.FieldsStayAtSizeHint)

        stepsize_entry = QLineEdit("1e-3")
        steps_entry = QLineEdit("80_000")
        posterior_samples_entry = QLineEdit("4000")

        parent.registerField("optimizer_stepsize", stepsize_entry)
        parent.registerField("optimizer_num_steps", steps_entry)
        parent.registerField("num_posterior_samples", posterior_samples_entry)

        bayesian_vi_layout.addRow("Optimizer stepsize:", stepsize_entry)
        bayesian_vi_layout.addRow("Number of optimization steps:", steps_entry)
        bayesian_vi_layout.addRow("Number of posterior samples:", posterior_samples_entry)

        config_page.addWidget(bayesian_vi_container)

        least_squares_container = QLabel("Please fill in initial values for varying parameters.")
        config_page.addWidget(least_squares_container)


class ModelPage(QWizardPage):
    def __init__(self, parent: QWizard):
        super().__init__(parent)

        self.setTitle("Configure Model Parameters")
        layout = QVBoxLayout(self)
        layout2 = QHBoxLayout()
        layout2.addWidget(ConstantsGroup(self))
        layout2.addWidget(SolverGroup(self))
        layout.addLayout(layout2)
        layout.addWidget(VariablesGroup(self))
        self.setCommitPage(True)

        parent.button(QWizard.WizardButton.CommitButton).clicked.connect(self.makeModel)

    def makeModel(self):
        self.wizard().model_parameters = lmfit.Parameters()

        for name, init, vary, min, max, val in zip(
            ["R1a", "R2a", "dwa", "R1b", "R2b", "kb", "fb", "dwb"],
            [
                "R1a_init_val",
                "R2a_init_val",
                "dwa_init_val",
                "R1b_init_val",
                "R2b_init_val",
                "kb_init_val",
                "fb_init_val",
                "dwb_init_val",
            ],
            [
                "R1a_vary",
                "R2a_vary",
                "dwa_vary",
                "R1b_vary",
                "R2b_vary",
                "kb_vary",
                "fb_vary",
                "dwb_vary",
            ],
            ["R1a_min", "R2a_min", "dwa_min", "R1b_min", "R2b_min", "kb_min", "fb_min", "dwb_min"],
            ["R1a_max", "R2a_max", "dwa_max", "R1b_max", "R2b_max", "kb_max", "fb_max", "dwb_max"],
            ["R1a_val", "R2a_val", "dwa_val", "R1b_val", "R2b_val", "kb_val", "fb_val", "dwb_val"],
        ):
            par_varies = not self.field(vary)  # field returns 0 if parameter varies...
            if par_varies:
                par_min = float(self.field(min))
                par_max = float(self.field(max))
            else:
                par_min = None
                par_max = None
            try:
                par_value = float(self.field(init) if par_varies else self.field(val))
            except ValueError:  # field probably empty. Choose average as init value
                par_value = (min + max) / 2

            self.wizard().model_parameters.add(name=name, value=par_value, vary=par_varies, min=par_min, max=par_max)


class FitPage(QWizardPage):
    def __init__(self, parent: QWizard):
        super().__init__(parent)

        self.setTitle("Performing Fitting...")
        layout = QVBoxLayout(self)
        btn = QPushButton("Press me to fit!")
        layout.addWidget(btn)
        btn.clicked.connect(self.perform_fit)

    def perform_fit(self):
        df = self.wizard().dataframe
        b0 = float(self.field("b0"))
        gamma = float(self.field("gamma")) * 2 * np.pi
        tp = float(self.field("tp"))
        powers = np.array([float(b1) for b1 in self.field("b1").split(",")])
        offsets = df.to_numpy(dtype=float).T[0]
        data = df.to_numpy(dtype=float).T[1:]
        model_parameters = self.wizard().model_parameters
        model_args = (offsets, powers, b0, gamma, tp)

        match self.field("solver"):
            case 0:  # "Symbolic"
                solver = batch_gen_spectrum_symbolic
            case 1:  # "Analytical":
                solver = batch_gen_spectrum_analytical
            case 2:  # "Numerical"
                solver = batch_gen_spectrum_numerical

        match self.field("fitting_method"):
            case 0:  # "Bayesian, MCMC"
                fitting_method = bayesian_mcmc
                num_warmup = int(n_warmup) if (n_warmup := self.field("num_warmup")) else None
                num_samples = int(n_samples) if (n_samples := self.field("num_samples")) else None
                num_chains = int(n_chains) if (n_chains := self.field("num_chains")) else None
                args = (model_parameters, model_args, data, solver, num_warmup, num_samples, num_chains)
            case 1:  # "Bayesian, ADVI"
                fitting_method = bayesian_vi
                optimizer_step_size = float(step_size) if (step_size := self.field("optimizer_stepsize")) else None
                optimizer_num_steps = int(n_steps) if (n_steps := self.field("optimizer_num_steps")) else None
                num_posterior_samples = int(n_samples) if (n_samples := self.field("num_posterior_samples")) else None
                args = (
                    model_parameters,
                    model_args,
                    data,
                    solver,
                    optimizer_step_size,
                    optimizer_num_steps,
                    num_posterior_samples,
                )
            case 2:  # "Nonlinear Least Squares"
                fitting_method = least_squares
                args = (model_parameters, model_args, data, solver)

        # Do this in a separate thread!
        self.result = fitting_method(*args)


class WizardApp(QWizard):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.setWindowTitle("Bloch-McConnellizer")
        self.setMinimumSize(QSize(940, 640))
        self.addPage(IntroPage())
        self.addPage(DataPage())
        self.addPage(ModelPage(self))
        self.addPage(FitPage(self))


def makeBold(widget: QWidget):
    f = widget.font()
    f.setWeight(QFont.Weight.DemiBold)
    widget.setFont(f)
    return widget


def write_qtable_to_df(table: QTableWidget):
    col_count = table.columnCount()
    row_count = table.rowCount()
    headers = [str(table.horizontalHeaderItem(i).text()) for i in range(col_count)]

    # df indexing is slow, so use lists
    df_list = []
    for row in range(row_count):
        df_list2 = []
        for col in range(col_count):
            table_item = table.item(row, col)
            df_list2.append(np.nan if table_item is None else str(table_item.text()))
        df_list.append(df_list2)

    df = pd.DataFrame(df_list, columns=headers).dropna(how="all")
    return df


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = WizardApp()
    window.show()
    sys.exit(app.exec())
